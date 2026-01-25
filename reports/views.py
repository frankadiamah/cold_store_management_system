
 
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from sales.models import Sale, SaleItem
from expenses.models import Expense
from inventory.models import Product
from django.http import HttpResponse,JsonResponse
import csv
import io
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from django.db.models import Sum
from datetime import datetime, timedelta
from users.utils import has_any_group
from django.template.loader import render_to_string

try:
    import openpyxl  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    OPENPYXL_AVAILABLE = True
except Exception:
    openpyxl = None
    # minimal fallback for get_column_letter to avoid NameError in static checks
    def get_column_letter(n):
        return str(n)
    OPENPYXL_AVAILABLE = False

 

 # WeasyPrint import attempt (may raise if system deps missing)
try:
    from weasyprint import HTML  # type: ignore
    WEASYPRINT_AVAILABLE = True
except Exception:
    WEASYPRINT_AVAILABLE = False

# xhtml2pdf fallback
try:
    from xhtml2pdf import pisa
    XHTML2PDF_AVAILABLE = True
except Exception:
    XHTML2PDF_AVAILABLE = False


# if you added CreditPayment model
try:
    from sales.models import CreditPayment
    HAS_CREDIT_PAYMENTS = True
except Exception:
    CreditPayment = None
    HAS_CREDIT_PAYMENTS = False


@login_required
@has_any_group( "SuperAdmin", "Admin","Accountant")
def summary(request):
    start = request.GET.get("start")
    end = request.GET.get("end")

    qs_sales = Sale.objects.all()
    qs_expenses = Expense.objects.all()

    if start:
        qs_sales = qs_sales.filter(timestamp__date__gte=start)
        qs_expenses = qs_expenses.filter(timestamp__date__gte=start)
    if end:
        qs_sales = qs_sales.filter(timestamp__date__lte=end)
        qs_expenses = qs_expenses.filter(timestamp__date__lte=end)

    total_sales = qs_sales.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")
    total_expenses = qs_expenses.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    gross_profit = total_sales - total_expenses

    # ✅ CREDIT METRICS
    credit_qs = qs_sales.filter(is_credit=True)

    credit_sales_total = credit_qs.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")
    # compute outstanding as sum of (total_amount - amount_paid)
    credit_outstanding_expr = ExpressionWrapper(F("total_amount") - F("amount_paid"), output_field=DecimalField())
    credit_outstanding = credit_qs.aggregate(total=Sum(credit_outstanding_expr))["total"] or Decimal("0.00")

    # option A (simple): use Sale.amount_paid
    credit_paid_via_sales = credit_qs.aggregate(total=Sum("amount_paid"))["total"] or Decimal("0.00")

    # option B (better): use CreditPayment records if you have them
    credit_paid_via_payments = Decimal("0.00")
    if HAS_CREDIT_PAYMENTS:
        pay_qs = CreditPayment.objects.filter(sale__in=credit_qs)
        credit_paid_via_payments = pay_qs.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    best_selling = (
        SaleItem.objects
        .values("product__name")
        .annotate(qty=Sum("quantity"))
        .order_by("-qty")[:10]
    )

    context = {
        "total_sales": total_sales,
        "total_expenses": total_expenses,
        "gross_profit": gross_profit,
        "best_selling": best_selling,

        # ✅ add these to template
        "credit_sales_total": credit_sales_total,
        "credit_outstanding": credit_outstanding,
        "credit_paid_via_sales": credit_paid_via_sales,
        "credit_paid_via_payments": credit_paid_via_payments,
        "start": start,
        "end": end,
    }
    return render(request, "reports/summary.html", context)

       
# @login_required
# @has_any_group("Admin","Accountant")
# def summary(request):
#     # date filters
#     start = request.GET.get("start")
#     end = request.GET.get("end")
#     qs_sales = Sale.objects.all()
#     qs_expenses = Expense.objects.all()
#     if start:
#         qs_sales = qs_sales.filter(timestamp__date__gte=start)
#         qs_expenses = qs_expenses.filter(timestamp__date__gte=start)
#     if end:
#         qs_sales = qs_sales.filter(timestamp__date__lte=end)
#         qs_expenses = qs_expenses.filter(timestamp__date__lte=end)

#     total_sales = qs_sales.aggregate(total=Sum("total_amount"))["total"] or 0
#     total_expenses = qs_expenses.aggregate(total=Sum("amount"))["total"] or 0

#     # approximate cost calculation: assume product.unit_price is sale price; to compute gross profit we need cost price.
#     # If you track cost price, use it. For demo we'll subtract expenses only.
#     gross_profit = total_sales - total_expenses

#     best_selling = SaleItem.objects.values("product__name").annotate(qty=Sum("quantity")).order_by("-qty")[:10]

#     context = {
#         "total_sales": total_sales,
#         "total_expenses": total_expenses,
#         "gross_profit": gross_profit,
#         "best_selling": best_selling,
#     }
#     return render(request, "reports/summary.html", context)

# @login_required
# def export_sales_csv(request):
#     sales = Sale.objects.all().order_by("-timestamp")
#     response = HttpResponse(content_type="text/csv")
#     response["Content-Disposition"] = f'attachment; filename="sales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
#     writer = csv.writer(response)
#     writer.writerow(["Sale ID", "Date", "Created By", "Total"])
#     for s in sales:
#         writer.writerow([s.id, s.timestamp, s.created_by.username if s.created_by else "", s.total_amount])
#     return response
# added by frank for exporting expenses to csv
@login_required
def export_expenses_csv(request):
    expenses = Expense.objects.all().order_by("-timestamp")
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="expenses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Expense ID", "Date", "Created By", "Amount", "Category", "Note"])
    for e in expenses:
        writer.writerow([e.id, e.timestamp, e.created_by.username if e.created_by else "", e.amount, e.category.name if e.category else "", e.note])
    return response





@login_required
@has_any_group("Admin","Accountant")
def export_sales_csv(request):
    sales = Sale.objects.all().order_by("-timestamp")
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="sales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Sale ID", "Date", "Created By", "Total"])
    for s in sales:
        writer.writerow([s.id, s.timestamp, s.created_by.username if s.created_by else "", s.total_amount])
    return response

@login_required
@has_any_group("SuperAdmin","Admin","Accountant")       
def export_sales_excel(request):
    if not OPENPYXL_AVAILABLE:
        return HttpResponse(
            "Excel export is unavailable because openpyxl is not installed. Install openpyxl to enable this feature.",
            status=500,
        )

    sales = Sale.objects.all().order_by("-timestamp")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    headers = ["Sale ID", "Date", "Created By", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for idx, s in enumerate(sales, start=2):
        ws.cell(row=idx, column=1, value=s.id)
        ws.cell(row=idx, column=2, value=s.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=idx, column=3, value=s.created_by.username if s.created_by else "")
        ws.cell(row=idx, column=4, value=float(s.total_amount))

    # Autosize columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="sales_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@login_required
@has_any_group("SuperAdmin","SubAdmin","Admin","Accountant")
def export_sales_pdf(request):
    sales = Sale.objects.all().order_by('-timestamp')

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 14)
    p.drawString(40, height - 50, "Sales Report")
    p.setFont("Helvetica", 9)
    p.drawString(40, height - 65, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    y = height - 90
    p.setFont("Helvetica-Bold", 9)
    p.drawString(40, y, "ID")
    p.drawString(90, y, "Date")
    p.drawString(220, y, "User")
    p.drawRightString(width - 40, y, "Total (₵)")
    p.line(35, y-3, width-35, y-3)
    y -= 14
    p.setFont("Helvetica", 9)

    for s in sales:
        if y < 60:
            p.showPage()
            y = height - 40
        p.drawString(40, y, str(s.id))
        p.drawString(90, y, s.timestamp.strftime("%Y-%m-%d"))
        p.drawString(220, y, s.created_by.get_full_name() if hasattr(s, 'created_by') and s.created_by else "")
        p.drawRightString(width - 40, y, f"{float(s.total_amount):.2f}")
        y -= 12

    p.showPage()
    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"'
    return response

 
def chart_sales_vs_expenses(request):
    # last 30 days
    days = int(request.GET.get("days", 30))
    labels = []
    sales_data = []
    expenses_data = []
    today = datetime.today().date()
    for i in range(days - 1, -1, -1):
        day = today - timedelta(days=i)
        labels.append(day.strftime("%Y-%m-%d"))
        ds = Sale.objects.filter(timestamp__date=day).aggregate(total=Sum("total_amount"))["total"] or 0
        de = Expense.objects.filter(timestamp__date=day).aggregate(total=Sum("amount"))["total"] or 0
        sales_data.append(float(ds))
        expenses_data.append(float(de))
    return JsonResponse({"labels": labels, "sales": sales_data, "expenses": expenses_data})
